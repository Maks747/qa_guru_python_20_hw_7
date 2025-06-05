import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR

def test_pdf():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('Список адресов.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'Список адресов' in text

def test_xlsx():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('Список адресов.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            value = sheet.cell(row=3, column=1).value
            name = 'Москва, ул.Осенняя'
            assert name in value

def test_csv():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('Список адресов в csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[2]
            assert second_row[0] == 'Moscow Osennyaya St.;15;5'

